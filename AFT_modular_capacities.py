# -*- coding: utf-8 -*-
"""
Created on Tue Nov 17 14:52:22 2020

@author: Shkurte Esati
-------------------------------------------------------------------------

Reads data for a flow thinning problem, builds the model, and solves it.

Input data:
    
set of IP nodes V
set of IP links E
set of optical nodes O
set of fibre links M
set of Demands D
volume traffic h(d)
alpha(e,s)
set of states S
set of sdmissible paths P(d)
set of links E(d,p) composing paths p ∈ P(d)
set of paths R(d,e) in P(d) that traverse link e
subset of states S(e), s ∈ S for which a particular link is not fully available
subset of links E(s) in E that are not fully available in state s ∈ S

Modeling variables:
Type 1 - aek      Number of modules of type k ∈ K that make up link e ∈ E  
Type 2 - xdp0     Nominal capacity of tunnel for demand d ∈ D routed over path p ∈ P(d) in failure free mode
Type 3 - xdps     Nominal capacity of tunnel for d ∈ D routed over path p ∈ P(d), when the network is in state s      
Type 4 - zdpe     -.-.-

Objective:
minimize sum(e ∈ E) aek * ck

Constraints:
Type 1:     forall IP links e ∈ E: sum(d ∈ D) sum(p ∈ p(d)) xdp0 <= sum (k ∈ K) aek * uk 
Type 2:     forall demands d ∈ D, forall states s ∈ S : sum(p ∈ p(d)) xdps >= h(d)
Type 3:     forall IP links e ∈ E, forall states s ∈ S(e): sum(d ∈ D) sum(p ∈ R(d,e)) xdps <= alpha(e,s) * sum (k ∈ K) aek * uk
Type 4:     forall demands d ∈ D, forall paths p ∈ p(d), forall states s ∈ S:  xdps <= xdp0
Type 5:     forall IP links e ∈ E, forall modules k ∈ K:  aek >= 0 aek ∈ Z
Type 6:     forall demands d ∈ D, forall paths p ∈ p(d): xdp0 >= 0
Type 7:     forall demands d ∈ D, forall paths p ∈ p(d), forall states s ∈ S:  xdps >= 0    
Type 8:     forall demands d ∈ D, forall paths p ∈ p(d), forall states s ∈ S:  zdpe >= 0

"""
from __future__ import print_function

import pandas as pd

import cplex


import os.path
from openpyxl import Workbook
import openpyxl




def main():
    
    flowthinning()
    
    
def flowthinning():
    
    #Read the data from excel files
    data = pd.ExcelFile(r'C:\Users\user\Desktop\Research_Project_Shkurte_Esati\Code\AFT_problem_modular_capacities\germanNetwork_inputdata.xlsx')
    
    c, var_typeInteger, var_type2, var_type3, var_AFT = configureproblem(data)

    c.solve()

    writeexcelfile(c, var_typeInteger, var_type2, var_type3, var_AFT)

def writeexcelfile(c, var_typeInteger, var_type2, var_type3, var_AFT):

    solution = c.solution

    # Print the solution status.
    print()
    print("Solution status = {0}: {1}".format(
        solution.get_status(),
        solution.get_status_string()))
    print()
    
    #################################################
    
    #                   Print Objective
    
    #################################################
    
    # Print te objective value
    print("Objective value = ", solution.get_objective_value())
    print()
    
    #################################################
    
    #                   Print Variables
    
    #################################################
    
    
    # num_of_variables gets the total number of variables
    num_of_variables = c.variables.get_num()
    
    print("Total number of variables is =  %d " % (num_of_variables))
    print()
    print()

    '''
    # prints the variables of type 1 with their corresponding values and name
    var_names = c.variables.get_names()
    x = solution.get_values(var_typeInteger)
    for j, val in enumerate(x):
        print("%d Variable {0} = %17.10g".format(var_names[j]) % (j, val))
    print()
    
    
    # prints the variables of type 2 with their corresponding values and name
    y = solution.get_values(var_type2)
    for j, val in enumerate(y):
        print("%d Variable {0} = %17.10g".format(var_names[j + len(x)]) % (j, val))
    print()    
    
    # prints the variables of type 3 with their corresponding values and name 
    z = solution.get_values(var_type3)
    for j, val in enumerate(z):
        print("%d Variable {0} = %17.10g".format(var_names[j + len(x) + len(y)]) % (j, val))    
    print()
    print()
    
     # prints the variables of type AFT with their corresponding values
    w = sum(var_AFT, [])
    var_affine = solution.get_values(w)
    for j, val in enumerate(var_affine):
        print("%d Variable {0} = %17.10g".format(var_names[j+ len(x) + len(y) + len(z)]) % (j+1, val))    
    print()
    print()
     
    #####################################################
    
    #                   Print Constraints in the screen
    
    #####################################################
    
    # num_of_constraints gets the total number of constraints
    num_of_constraints = c.linear_constraints.get_num()
    print("Total number of contraints is =  %d " % (num_of_constraints))
    print()
    print()
    
    # constraint_names gets the names of constraints
    constraint_names = c.linear_constraints.get_names()
    
    # c1_names gets the names of constraint type 1 in a list
    c1_names = [s for s in constraint_names if 'c1_' in s]
    print("Constraints of type 1 are: ") 
    print(*c1_names, sep = ", ")
    print()
    
    # c2_names gets the names of constraint type 2 in a list
    c2_names = [s for s in constraint_names if 'c2_' in s]
    print("Constraints of type 2 are: ")
    print(*c2_names, sep = ", ")
    print()
    
    
    # c3_names gets the names of constraint type 3 in a list
    c3_names = [s for s in constraint_names if 'c3_' in s]
    print("Constraints of type 3 are: ")
    print(*c3_names, sep = ", ")
    print() 
    
    # c4_names gets the names of constraint type AFT in a list
    c4_names = [s for s in constraint_names if 'c4_' in s]
    print("Constraints of type AFT are: ")
    print(*c4_names, sep = ", ")
    print()
    '''

    ############################################################################################
    
    #                  Print variables and constraints in Excel file
    
    ############################################################################################
    
    fname = r'Results_AffineFlowThinning.xlsx'
    
    if os.path.isfile(fname):
        
        print('old file')
        
        book = openpyxl.load_workbook(fname)
        
        book.remove(book['Objective'])
        book.remove(book['Var type Integer'])
        book.remove(book['Var Type 2'])
        book.remove(book['Var Type 3'])
        book.remove(book['Var Type AFT'])
        book.remove(book['Constraints'])
        book.remove(book['Verification of Constraints'])
        
    else:
        
        print('new file')
        
        book = Workbook()
        
    book.create_sheet('Objective')
    book.create_sheet('Var type Integer')
    book.create_sheet('Var Type 2') 
    book.create_sheet('Var Type 3')
    book.create_sheet('Var Type AFT')
    book.create_sheet('Constraints')
    book.create_sheet('Verification of Constraints')
    
    ######################################################################
    
    #                    Print Objective
    
    
    ######################################################################
    
    sheet = book['Objective']
    sheet['A1'] = 'Objective Value:'
    
    sheet['B1'] = solution.get_objective_value()
    
    
    ######################################################################
    
    #                    Print Variables of Type 1
    
    
    ######################################################################
    
    sheet = book['Var type Integer']
    sheet['A1'] = 'Index:'
    sheet['B1'] = 'Name of variable:'
    sheet['C1'] = 'Value:'
    
    var_names = c.variables.get_names()
    x = solution.get_values(var_typeInteger)
    for j,val in enumerate(x):
        
        sheet['A' + str(j+2)] = j
        sheet['B' + str(j+2)] = var_names[j]
        sheet['C' + str(j+2)] = val
       
        
    ######################################################################
    
    #                    Print Variables of Type 2
    
    
    ######################################################################
    
    sheet = book['Var Type 2']
    sheet['A1'] = 'Index:'
    sheet['B1'] = 'Name of variable:'
    sheet['C1'] = 'Value [Gbit/s]:'
    
    y = solution.get_values(var_type2)
    for j,val in enumerate(y):
        
        sheet['A' + str(j+2)] = j
        sheet['B' + str(j+2)] = var_names[j+len(x)]
        sheet['C' + str(j+2)] = val
           
    ######################################################################
    
    #                  Print Variables of Type 3
    
    
    ######################################################################
    
    sheet = book['Var Type 3']
    sheet['A1'] = 'Index:'
    sheet['B1'] = 'Name of variable:'
    sheet['C1'] = 'Value [Gbit/s]:'
    
    z = solution.get_values(var_type3)
    for j,val in enumerate(z):
        
        sheet['A' + str(j+2)] = j
        sheet['B' + str(j+2)] = var_names[j+len(x)+len(y)]
        sheet['C' + str(j+2)] = val
               

    ######################################################################
    
    #                  Print Variables of Type AFT
    
    
    ######################################################################
    
    sheet = book['Var Type AFT']
    sheet['A1'] = 'Index:'
    sheet['B1'] = 'Name of variable:'
    sheet['C1'] = 'Value [Gbit/s]:'
    
    w = sum(var_AFT, [])
    var_affine = solution.get_values(w)
    for j,val in enumerate(var_affine):
        
        sheet['A' + str(j+2)] = j
        sheet['B' + str(j+2)] = var_names[j+len(x)+len(y)+len(z)]
        sheet['C' + str(j+2)] = val

    ######################################################################

    #                    Print Constraints 

    ######################################################################

 

    con_names   = c.linear_constraints.get_names()
    con_number  = c.linear_constraints.get_num()
    con_rows    = c.linear_constraints.get_rows()
    con_rhs     = c.linear_constraints.get_rhs()
    con_senses  = c.linear_constraints.get_senses()
    var_values = solution.get_values()

    sheet =  book['Constraints']

    sheet['A1'] = 'Overview over all constraints'


    for i in range(con_number):
        
        con_string = con_names[i] + ':  '
        spair = con_rows[i]
        ind, val = spair.unpack()

        for j in range(len(ind)):

                if val[j] >= 0:

                        con_string = con_string + '+ ' + str(val[j]) + ' * ' + var_names[ind[j]] + ' '

                elif val[j] < 0:

                        con_string = con_string + str(val[j]) + ' * ' + var_names[ind[j]] + ' '

        if con_senses[i] == 'G':

                con_string = con_string + '>= '

        elif con_senses[i] == 'L':

                con_string = con_string + '<= '

        elif con_senses[i] == 'E':

                con_string = con_string + '= '

        elif con_senses[i] == 'R':

                con_string = con_string + 'is in the range of: '

        con_string = con_string + str(con_rhs[i])

        sheet['A' + str(i+2)] = con_string

    ################
    sheet =  book['Verification of Constraints']
    sheet['A1'] = 'Overview over all constraints:'
    sheet['O1'] = 'Left Handside of the constraints:'
    
    for i in range(con_number):

        con_values = 0        
        con_string = con_names[i] + ':  '
        spair = con_rows[i]
        ind, val = spair.unpack()

        for j in range(len(ind)):

                if val[j] >= 0:

                        con_string = con_string + '+ ' + str(val[j]) + ' * ' + str(var_values[ind[j]])  + ' '
                        con_values = con_values + val[j]  *  var_values[ind[j]] 
                        
                elif val[j] < 0:

                        con_string = con_string + str(val[j]) + ' * ' + str(var_values[ind[j]])  + ' '
                        con_values = con_values 
                        
        if con_senses[i] == 'G':

                con_string = con_string + '>= '
                
        elif con_senses[i] == 'L':

                con_string = con_string + '<= '
                
        elif con_senses[i] == 'E':

                con_string = con_string + '= '
                
        elif con_senses[i] == 'R':

                con_string = con_string + 'is in the range of: '
                
        con_string = con_string + str(con_rhs[i])
        
        sheet['A' + str(i+2)] = con_string
        sheet['O' + str(i+2)] = con_values
        
    book.save(fname)
    
def configureproblem(data):
    
    ##############################################################################
    
    #                   create a CPLEX object 
    
    ##############################################################################

    c = cplex.Cplex()
    
    CPLX_LP_PARAMETERS = {
    'simplex.tolerances.optimality' : 1e-9,
    'simplex.tolerances.feasibility' : 1e-9
    } 

    c.parameters.simplex.tolerances.optimality.set(CPLX_LP_PARAMETERS['simplex.tolerances.optimality'])
    c.parameters.simplex.tolerances.feasibility.set(CPLX_LP_PARAMETERS['simplex.tolerances.feasibility'])
    


    ##############################################################################
    
    #                     Initialize input data 
    
    ##############################################################################
    
    
    #Here we read the data from excel sheets and columns
    
    #data specifies the path for the excel file
    #each set will be read out of the excel file with its specific column
 
    set_IPLinks = pd.read_excel(data, sheet_name='IPlinks' , usecols= ['IP links E'])
    set_Demands = pd.read_excel(data, sheet_name='Demands' , usecols= ['D Demands'])
    set_TrafficVolume = pd.read_excel(data, sheet_name='TrafficVolume' , usecols= ['h(d)'])
    set_States = pd.read_excel(data, sheet_name='States', usecols=['States S'])    
    set_Paths = pd.read_excel(data, sheet_name='Paths', usecols=['Paths P(d)'])
    set_paths_per_demand = pd.read_excel(data, sheet_name='Paths', usecols=['paths per demand'])
    set_Paths_Total = pd.read_excel(data, sheet_name = "Paths")
    subset_states = pd.read_excel(data, sheet_name='subset states', usecols= ['Subset of States S€'])
    set_paths_per_link = pd.read_excel(data, sheet_name='Paths R(d,e)', usecols= ['Paths per link'])
    alpha = pd.read_excel(data, sheet_name='Alpha')
    set_modules = pd.read_excel(data, sheet_name='Modules', usecols= ['capacity per module uk'])
    costs_per_module = pd.read_excel(data, sheet_name='CostsPerModule', usecols=['Set of costs C of modules K'])
    set_links_per_path = pd.read_excel(data, sheet_name='Links E(d,p)', usecols=['links E(d,p)'])
    
    
    # set_paths_per_demand and set_paths_per_demand1 access different sheets, but they take the same data namly Paths per demand
    # They are used in different costraints, according to the sheet they access.
    set_paths_per_demand = pd.read_excel(data, sheet_name='Paths', usecols=['paths per demand'])
    set_paths_p_demand1 = pd.read_excel(data, sheet_name='Links E(d,p)', usecols=['Paths P(d)'])
    
    # Calc column length
    N_IPlinks = len(set_IPLinks)
    N_Demands = len(set_Demands)
    N_States = len(set_States)
    N_Paths = len(set_Paths)
    N_Modules = len(set_modules)
    
    #gets the values from cost dataframe and makes it a list
    cost = costs_per_module["Set of costs C of modules K"].tolist()
    modules = set_modules["capacity per module uk"].tolist()
    capacity_permodule  = [float(i) for i in modules]
    
    ######################################################################################################
    
    # Create a list with the number of paths belonging to each demand
    # numPaths_of_d variable will contain the number of paths for each specific demands saved into a list
    # we need this variable later to loop through the paths of each demand 
    # Note: Every demand has a different number of paths
    #######################################################################################################

    # extract column from pandas dataframe 
    demandList = set_Paths_Total["D Demands"].tolist()


    # loop through all entries
    counter = 0
    numPaths_of_d = []
    loop_counter = 0

    for d in demandList:
        
            loop_counter = loop_counter + 1;  
            if isinstance(d,str) == 0:
                    counter = counter + 1
                    if (loop_counter == len(demandList)): 
                            numPaths_of_d.append(counter)
                
                
            else:
                    if loop_counter == 1:
                            counter = 1
                        
                    else:
                        numPaths_of_d.append(counter)
                        counter = 1
                        
                        
    ##############################################################################
    
    #             Add Objective with the purpose of minimizing it 
    
    #                minimize sum(e ∈ E) Ye0  * cost(e)
    
    ##############################################################################
    
    c.objective.set_sense(c.objective.sense.minimize)
    c.objective.set_name("minimze nominal capacity at failure free mode")
    
    
    
    #########################################################################################################################################
    
    #                 Add variables
    
    # Note: CPLEX assigns each created variable a specific unique index, which allows an easy selection of the needed variable.
    
    # to add variables with CPLEX we need these arguments: 
    #1. objective - if this variable is used in the objective - normally is a list of ones whose length is equal to the number of variables
    #2. lb - is the lower bound of the variables. It is usually 0.
    #3. ub -is the upper bound of the variables. It is usually infinity.
    #4. types - specify the type of variables. In our case all the variables are of type continuous
    #5. names (optional) - specifies the names of variables
    
    ########################################################################################################################################
    
    
    
    ######################################################################################
    
    #                     Variable of Type 1 - aek  
    
    #            capacity of IP link e ∈ E in failure free mode
    
    #                forall IP links e ∈ E:  aek >= 0
    #            Total number of this variable is equal to the number of IP links.
    
    ######################################################################################
    
   
    #variable of type 1 is used together with cost in the objective, therefore in the argument obj we use also cost.
    varnames_typeInteger = ["ae" + str(e + 1) + "k" + str(k + 1) for e in range(N_IPlinks) for k in range(N_Modules)]
    var_typeInteger = list(c.variables.add(obj = cost * N_IPlinks, 
                                lb=[0.0] * N_IPlinks * N_Modules,
                                ub=[cplex.infinity] * N_IPlinks * N_Modules,
                                types= [c.variables.type.integer]* N_IPlinks * N_Modules,
                                names = varnames_typeInteger
                                ))
    
    
    ###########################################################################################################################
    
    #                                        Variable of Type 2 - xdp0     
    
    #              Nominal capacity of tunnel for demand d ∈ D routed over path p ∈ P(d) in failure free mode
    #                 forall demands d ∈ D, forall paths p ∈ p(d):    xdp0 >= 0
    #                      Total number of this variable is: sum(d ∈ D) |P(d)|
    
    ###########################################################################################################################
   
     
    
    varnames_type2 = ["xd" + str(d + 1) + "p" + str(p + 1)  for d in range(N_Demands) for p in range(numPaths_of_d[d])]    
    var_type2 = list(c.variables.add(lb=[0.0] * N_Paths,
                                ub=[cplex.infinity] * N_Paths,
                                types= [c.variables.type.continuous] * N_Paths,
                                names = varnames_type2
                                ))
    
    
    ###############################################################################################################################
    
    #                                      Variable of Type 3 - xdps   
    
    #               Nominal capacity of tunnel for d ∈ D routed over path p ∈ P(d), when the network is in state s 
    #               forall demands d ∈ D, forall paths p ∈ p(d), forall states s ∈ S:  xdps >= 0
    #                   Total number of this variable is: |S| * sum(d ∈ D) |P(d)|
    
    ###############################################################################################################################
    
    
    varnames_type3 = ["xd" + str(d + 1) + "p" + str(p + 1) + "s" + str(s + 1) for s in range(N_States) for d in range(N_Demands) for p in range(numPaths_of_d[d])]
    var_type3 = list(c.variables.add(lb=[0.0] * N_Paths * N_States,
                                ub=[cplex.infinity] * N_Paths * N_States,
                                types= [c.variables.type.continuous]* N_Paths * N_States,
                                names = varnames_type3
                                ))

    ###############################################################################################################################
    
    #                                      Variable of Type 4 - zdpe   
    
    #                 forall demands d ∈ D, forall paths p ∈ p(d), forall states s ∈ S:  zdpe >= 0              
    #                       Total number of this variable is:  sum(d ∈ D) sum(p ∈ P(d)) |E(d)|
    
    ###############################################################################################################################
    
    #empty list to save variables of type AFT
    var_AFT = []
    counter = 0
    for d in range(N_Demands): 
        
        for n in range(numPaths_of_d[d]):

            #we need to iterate over links per path
            # In links_per_path we save the elements of the links that go through paths of demands - for example: e1,e2
            links_per_path = set_links_per_path.at[counter, 'links E(d,p)']
            
            # In newstring_forLinks we save the numbers of the elements of links - for example: 1,2
            newstring_forLinks = ''.join((ch if ch in '0123456789.' else ' ') for ch in links_per_path)
            
            # in listOfNumbers_forLinks we save the elements of newstr_forLinks as a list - for example: [1,2]
            listOfNumbers_forLinks = [int(i) for i in newstring_forLinks.split()] 
            
            # set_paths_per_demand and set_paths_per_demand1 access different sheets, but they take the same data namly Paths per demand
            # Here we need to loop through each path of demands
            
            # In paths_per_demand we save the elements of paths per demand - for example: p1,p2
            paths_per_demand = set_paths_p_demand1.at[counter, 'Paths P(d)']
            
            # In newstring_forPaths we save the numbers of the elements of paths per demand - for example: 1,2
            newstring_forPaths = ''.join((ch if ch in '0123456789.' else ' ') for ch in paths_per_demand)
            
            # in listOfNumbers_forPaths we save the elements of newstr_forPaths as a list - for example: [1,2]
            listOfNumbers_forPaths = [int(i) for i in newstring_forPaths.split()]
              
            #variable names for type AFT        
            varnames_typeAFT = ["zd" + str(d + 1) + "p" + str(p) + "e" + str(e) for e in listOfNumbers_forLinks for p in listOfNumbers_forPaths]
            
            counter = counter + 1
    
            var_listAFT = list(c.variables.add(lb=[0.0] * len(listOfNumbers_forLinks),
                                        ub=[cplex.infinity] * len(listOfNumbers_forLinks),
                                        types= [c.variables.type.continuous] * len(listOfNumbers_forLinks),
                                        names = varnames_typeAFT
                                        ))
            
            # var_AFT variable will be populated with elements of var_listAFT
            var_AFT.append(var_listAFT) 

    
    ##########################################################################
    
    #                                   Constraints
    
    # Remark: When a constraint has a variable on the right hand side or a variable multiplied with a parameter,
    # we need to bring it to the left hand side, otherwise cplex doesnt recognize it as a variable and only gets the index of it.
    # All constraints for FT and AFT problem formulation are linear constraints.
    
    # to add constraints with cplex we need these arguments: 
    #1. lin_expr - is a matrix in list-of-lists format. lin_expr contains: ind and val as arguments.
    #   ind - here we specify the variable type that is needed for the current constraint. Because CPLEX assigns unique index to variables, we can access then using these indexes.
    #   val - here we specify the coefficients in front of variable indicies
    #2. senses - specifies the senses of the linear constraint. We use these types:
    #   - L for less-than
    #   - G for greater-than
    #   - E for equal
    #3. rhs - is the right hand side of the equation. Most of the time rhs is a list of zeros since the variables that are on the right hand side of the equation, we can bring them to the left hand side of the equation.
    #5. names (optional) - specifies the names of costraints.
    
    ##########################################################################
    
    
    
    ##############################################################################
    
    #            Constraint of type 1: Capacity constraint
    #
    #                       forall IP links e ∈ E: 
    
    #           sum(d ∈ D) sum(p ∈ p(d)) xdp0 - sum (k ∈ K) aek * uk  <= 0
    #           Total number of constraints of type 1: |E|
    
    ###############################################################################
    
    
    for e in range(N_IPlinks):
            
            # names of the constraints
            constnames_type1 = ["c1_" + str(e)]
            
            # right hand side is a list of zeros since the variables of type one we can bring them to the left hand side
            rhs1 = [0.0]
            
            # In set_of_p_perlink we save the elements of the set of Paths per link - for example: p0,p3,p4,p7,p9,p11
            set_of_p_perlink = set_paths_per_link.at[e,'Paths per link']
            
            # In newstr we save the numbers of the elements of the set of Paths per link - for example: 0,3,4,7,9,11
            newstr = ''.join((ch if ch in '0123456789.-e' else ' ') for ch in set_of_p_perlink)
            
            # in listOfNumbers we save the elements of newstr as a list - for example: [0,3,4,7,9,11]
            listOfNumbers = [int(i) for i in newstr.split()]
 
            # indicies get the variables 
            # we use the list listOfNumbers to loop through the elements of variable of type 2
            ind = [var_type2[p] for p in listOfNumbers] + [var_typeInteger[k + e*N_Modules] for k in range(N_Modules)]
            
            #we multiply each element of capacity_permodule with -1 and we save it in capacity_permodule_negative
            capacity_permodule_negative = [element * (-1) for element in capacity_permodule]
            
            # val gives the coefficients of the indicies, length of the list is the same as the length of listOfNumbers
            # capacity_permodule_negative contains the coefficients of var_typeInteger
            val = [1.0] * len(listOfNumbers) + capacity_permodule_negative         
            
            rows = [[ind, val]]
            
            # lin_expr is a matrix in list-of-lists format.
            # senses specifies the senses of the linear constraint -L means less-than
            c.linear_constraints.add(lin_expr = rows,
                                     senses="L", 
                                     rhs=rhs1,
                                     names = constnames_type1)

    ######################################################################################
    
    #               Constraint of type 2: Flow conservation constraint
    
    #                 forall demands d ∈ D, forall states s ∈ S : 
    
    #                    sum(p ∈ p(d)) xdps >= h(d)
    #                 Total number of constraints of type 2: |D|* (|S| + 1)
    
    ######################################################################################
    
    
    for d in range(N_Demands):  
        
        # In traffic_volume are saved the elements of the traffic matrix
        traffic_volume = set_TrafficVolume.at[d,'h(d)']
        
        # In set_p_per_demand we save the elements of the set of Paths per demand - for example: p0,p1
        set_p_per_demand = set_paths_per_demand.at[d, 'paths per demand']
        
        # In newstr we save the numbers of the elements of the set of Paths per demand - for example: 0,1
        newstr = ''.join((ch if ch in '0123456789.-e' else ' ') for ch in set_p_per_demand)
        
        # in listOfNumbers we save the elements of newstr as a list - for example: [0,1]
        listOfNumbers = [int(i) for i in newstr.split()]
        
        # constraint of type 2 contains all the states from the set of States plus the state 0 - state of network in failure free mode
        for s in range(N_States + 1):
            
            # names of the constraints
            constnames_type2 = ["c2_" + str(s + (N_States + 1) * d)]
            
            # to write the contraints of type 2 for state 0 - failure free mode, we need to get the variables of type2 - xdp0 
            
            if (s == 0):
                
                # right hand side is a list of elements of the traffic matrix
                rhs1 = [float(traffic_volume)]
                
                # indicies get the variables
                # we use the list listOfNumbers to loop through the elements of variable of type 2
                ind = [var_type2[k] for k in listOfNumbers]
                
                # val gives the coefficients of the indicies, length of the list is the same as the length of listOfNumbers
                val = [1.0] * len(listOfNumbers)
                
                rows = [[ind, val]]
                
                # lin_expr is a matrix in list-of-lists format.
                # senses specifies the senses of the linear constraint -G means greater-than
                c.linear_constraints.add(lin_expr = rows,
                                         senses="G", 
                                         rhs=rhs1,
                                         names = constnames_type2)
            
            # to write the contraints of type 2 for all other states in the set of States S, then we need to get the variables of type3 - xdps 
            else:
                
                rhs1 = [float(traffic_volume)]
                
                # indicies get the variables
                # we use the list listOfNumbers to loop through the elements of variable of type 3
                ind = [var_type3[k + (s - 1) * N_Paths] for k in listOfNumbers]
                
                val = [1.0] * len(listOfNumbers)
                
                rows = [[ind, val]]
                
                c.linear_constraints.add(lin_expr = rows,
                                         senses="G", 
                                         rhs=rhs1,
                                         names = constnames_type2)   
                
                
    ############################################################################################################
    
    #                      Constraint of type 3: Capacity constraint 
    
    #                     forall IP links e ∈ E, forall states s ∈ S(e): 
    
    #                sum(d ∈ D) sum(p ∈ R(d,e)) xdps - alpha(e,s) * sum (k ∈ K) aek * uk <=  0 

    #                 Total number of constraints of type 3: sum(e ∈ E) |S(e)|

    ############################################################################################################
    
    #counter needed to name the constraints
    cnt_link_states = 0
    for e in range(N_IPlinks):
        
        # In subset_of_states we save the elements of the subset of states per link - for example: s1,s2
        subset_of_states = subset_states.at[e,'Subset of States S€']
        
        # In newstr we save the numbers of the elements of the subset of states per link - for example: 1,2 
        newstr = ''.join((ch if ch in '0123456789.-e' else ' ') for ch in subset_of_states)
        
        # in listOfNumbers we save the elements of newstr as a list - for example: [1,2]
        listOfNumbers = [int(i) for i in newstr.split()]
        
        #counter needed to name the constraints
        ind_naming = 0
        
        for s in (listOfNumbers):
            
            # names of the constraints
            constnames_type3 = ["c3_" + str(ind_naming + cnt_link_states)]
            
            #counter increased by 1
            ind_naming = ind_naming + 1
            
            # right hand side is a list of zeros since the variables of type one we can bring them to the left hand side
            rhs1 = [0.0]
            
            # In set_of_p_perlink we save the elements of the set of Paths per link - for example: p0,p3,p4,p7,p9,p11
            set_of_p_perlink = set_paths_per_link.at[e,'Paths per link']
            
            # In newstr1 we save the numbers of the elements of the set of Paths per demand - for example: 0,3,4,7,9,11
            newstr1 = ''.join((ch if ch in '0123456789.-e' else ' ') for ch in set_of_p_perlink)
            
            # in listOfNumbers we save the elements of newstr as a list - for example: [0,3,4,7,9,11]
            listOfNumbers1 = [int(i) for i in newstr1.split()]
            
            # indicies get the variables
            # we use the list listOfNumbers1 to loop through the elements of variable of type 3
            ind = [var_type3[k + (s - 1) * N_Paths] for k in listOfNumbers1] + [var_typeInteger[k + e*N_Modules] for k in range(N_Modules)]
            
            # val gives the coefficients of the indicies, length of the list is the same as the length of listOfNumbers1
            # alpha.iloc[e,s] gets elements as a matrix from excel sheet alpha
            # we multiply each element of capacity_permodule with -1 and the corresponding alpha value, we save the result of this multiplication in result
            result = [element * (- float(alpha.iloc[e,s])) for element in capacity_permodule] 
            
            # result contains the coefficients of var_typeInteger
            val = [1.0] * len(listOfNumbers1) + result
            
            rows = [[ind, val]]
            
            # lin_expr is a matrix in list-of-lists format.
            # senses specifies the senses of the linear constraint -L means less-than
            c.linear_constraints.add(lin_expr = rows,
                                     senses="L", 
                                     rhs=rhs1,
                                     names = constnames_type3) 
        
        # the counter is increased by the length of the list listOfNumbers
        cnt_link_states = cnt_link_states + len(listOfNumbers)
    
    ############################################################################################################
    
    #                      Constraint of type 4: AFT constraint 
    
    #                     forall states s ∈ S, forall demands d ∈ D, forall paths p ∈ p(d): 
    
    #                    xdps - xdp0 + sum(e ∈ E(d,p)) B(e,s) * zdpe = 0
    
    #                 Total number of constraints of type 4: |S| * sum(d ∈ D) |P(d)|
    
    ############################################################################################################
    
    
    for s in range(N_States):
    
        for d in range(N_Demands): 
            
            # In set_p_per_demand we save the elements of the set of Paths per demand - for example: p0,p1
            set_p_per_demand = set_paths_per_demand.at[d, 'paths per demand']
            
            # In newstr we save the numbers of the elements of the set of Paths per demand - for example: 0,1
            newstr = ''.join((ch if ch in '0123456789.-e' else ' ') for ch in set_p_per_demand)
            
            # in listOfNumbers we save the elements of newstr as a list - for example: [0,1]
            listOfNumbers = [int(i) for i in newstr.split()]
            
            for p in listOfNumbers:
                
                # names of the constraints
                constnames_typeAFT = ["c4_" + str(p + N_Paths * s)]
                
                # right hand side is a list of zeros since the variables of type one we can bring them to the left hand side
                rhs1 = [0.0]
                
                #we need to iterate over links per path
                # In links_per_path we save the elements of the links that go through paths of demands - for example: e1,e2
                links_per_path = set_links_per_path.at[p, 'links E(d,p)']
            
                # In newstring_forLinks we save the numbers of the elements of links - for example: 1,2
                newstring_forLinks = ''.join((ch if ch in '0123456789.' else ' ') for ch in links_per_path)
            
                # in listOfNumbers_forLinks we save the elements of newstr_forLinks as a list - for example: [1,2]
                listOfNumbers_forLinks = [int(i) for i in newstring_forLinks.split()] 
                                
                # indicies get the variables
                # we use the list listOfNumbers1 to loop through the elements of variable of type 3
                ind = [var_type3[p + N_Paths * s]] + [var_type2[p]] + var_AFT[p]
                
                # val gives the coefficients of the indicies
                val = [1.0] + [-1.0] 
                
                #an extra loop to get the coefficient beta depending on how many variables AFT we need in this constraint
                #Note: 1 - alpha.iloc[e - 1  ,s + 1] = [beta.iloc[e - 1 ,s + 1]]
                for e in listOfNumbers_forLinks:
                
                    val = val + [float(1 - alpha.iloc[e - 1  ,s + 1])]
                
                rows = [[ind, val]]
                
                # lin_expr is a matrix in list-of-lists format.
                # senses specifies the senses of the linear constraint -L means less-than
                c.linear_constraints.add(lin_expr= rows,
                                                   senses="E", 
                                                   rhs= rhs1,
                                                   names = constnames_typeAFT)
    

    
    
    return c, var_typeInteger, var_type2, var_type3, var_AFT
    
    
if __name__ == "__main__":
    main()
    

